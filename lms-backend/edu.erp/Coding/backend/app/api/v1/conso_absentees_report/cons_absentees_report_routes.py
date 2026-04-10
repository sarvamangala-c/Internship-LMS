from typing import Optional
from fastapi import APIRouter, Depends, Header
from sqlalchemy.orm import Session
from sqlalchemy import text, bindparam, or_
from app.core.database import get_db
from app.utils.http_return_helper import returnSuccess, returnException
from .cons_absentees_report_schema import *


print("DEBUG: Loading Consolidated Absentees Report Module...")
router = APIRouter(tags=["Consolidated Absentees Report"])

# --------------------------------------------------
# 1. DROPDOWNS
# --------------------------------------------------

@router.get("/departments")
@router.post("/departments")
def get_departments(db: Session = Depends(get_db), org_id: Optional[int] = Header(None)):
    print(f"DEBUG: get_departments called with org_id={org_id}")
    try:
        org = org_id or 1
        # Re-enabling org_id filter but allowing NULL for safety
        result = db.execute(text("""
            SELECT 
                dept_id AS id,
                dept_name AS name,
                dept_id AS department_id,
                dept_name AS dept_name
            FROM iems_department
            WHERE (org_id = :org OR org_id IS NULL)
        """), {"org": org}).fetchall()

        print(f"DEBUG: Found {len(result)} departments in DB")
        return returnSuccess([dict(row._mapping) for row in result])
    except Exception as e:
        print(f"DEBUG: get_departments error: {str(e)}")
        return returnException(str(e))


@router.get("/programs/{department_id}")
@router.post("/programs")
def get_programs(department_id: Optional[int] = None, payload: Optional[ProgramRequest] = None, db: Session = Depends(get_db), org_id: Optional[int] = Header(None)):
    try:
        dept = department_id or (payload.department_id if payload else None)
        org = org_id or 1
        
        query = """
            SELECT 
                pgm_id AS id, 
                pgm_title AS name,
                pgm_id AS program_id,
                pgm_title AS pgm_title
            FROM iems_program 
            WHERE (org_id = :org OR org_id IS NULL)
        """
        params = {"org": org}
        
        if dept and int(dept) > 0:
            query += " AND dept_id = :dept"
            params["dept"] = dept

        result = db.execute(text(query), params).fetchall()
        print(f"DEBUG: Found {len(result)} programs for dept={dept}")
        return returnSuccess([dict(row._mapping) for row in result])
    except Exception as e:
        return returnException(str(e))


@router.get("/curriculum/{program_id}")
@router.post("/curriculum")
def get_curriculum(program_id: Optional[int] = None, payload: Optional[CurriculumRequest] = None, db: Session = Depends(get_db), org_id: Optional[int] = Header(None)):
    try:
        prog = program_id or (payload.program_id if payload else None)
        org = org_id or 1
        
        query = """
            SELECT 
                academic_batch_id AS id, 
                academic_batch_desc AS name,
                academic_batch_id AS academic_batch_id,
                academic_batch_desc AS academic_batch_desc
            FROM iems_academic_batch 
            WHERE (org_id = :org OR org_id IS NULL)
        """
        params = {"org": org}
        
        if prog and int(prog) > 0:
            query += " AND pgm_id = :prog"
            params["prog"] = prog

        result = db.execute(text(query), params).fetchall()
        print(f"DEBUG: Found {len(result)} curriculum/batches for prog={prog}")
        return returnSuccess([dict(row._mapping) for row in result])
    except Exception as e:
        return returnException(str(e))


@router.get("/terms/{curriculum_id}")
@router.post("/terms")
def get_terms(curriculum_id: Optional[int] = None, payload: Optional[TermRequest] = None, db: Session = Depends(get_db), org_id: Optional[int] = Header(None)):
    try:
        cur = curriculum_id or (payload.curriculum_id if payload else None)
        org = org_id or 1
        
        query = """
            SELECT 
                semester_id AS id, 
                semester_desc AS name,
                semester_id AS semester_id,
                semester_desc AS semester_desc
            FROM iems_semester 
            WHERE (org_id = :org OR org_id IS NULL)
        """
        params = {"org": org}
        
        if cur and int(cur) > 0:
            query += " AND academic_batch_id = :cur"
            params["cur"] = cur

        result = db.execute(text(query), params).fetchall()
        print(f"DEBUG: Found {len(result)} terms for curriculum={cur}")
        return returnSuccess([dict(row._mapping) for row in result])
    except Exception as e:
        return returnException(str(e))


@router.get("/sections/{semester_id}")
@router.post("/sections")
def get_sections(semester_id: Optional[int] = None, payload: Optional[SectionRequest] = None, db: Session = Depends(get_db), org_id: Optional[int] = Header(None)):
    try:
        term = semester_id or (payload.semester_id if payload else None)
        org = org_id or 1
        
        query = """
            SELECT 
                id AS id, 
                section AS name,
                id AS section_id,
                section AS section_name
            FROM iems_section 
            WHERE (org_id = :org OR org_id IS NULL)
        """
        params = {"org": org}
        
        if term and int(term) > 0:
            query += " AND semester_id = :term"
            params["term"] = term

        result = db.execute(text(query), params).fetchall()
        print(f"DEBUG: Found {len(result)} sections for term={term}")
        return returnSuccess([dict(row._mapping) for row in result])
    except Exception as e:
        return returnException(str(e))

# --------------------------------------------------
# 2. DEFAULT DATE API
# --------------------------------------------------

@router.get("/date-info")
def get_date_info(db: Session = Depends(get_db)):
    try:
        latest_date = db.execute(text("""
            SELECT MAX(attendance_date) as latest_date
            FROM lms_manage_attendance
        """)).scalar()

        scheduled_dates = db.execute(text("""
            SELECT DISTINCT attendance_date
            FROM lms_manage_attendance
            ORDER BY attendance_date DESC
            LIMIT 30
        """)).fetchall()

        return returnSuccess({
            "latest_attendance_date": latest_date,
            "scheduled_dates": [row[0] for row in scheduled_dates]
        })
    except Exception as e:
        return returnException(str(e))


# --------------------------------------------------
# 3. MAIN REPORT
# --------------------------------------------------

@router.get("/report")
def test_report_get():
    print("DEBUG: test_report_get called")
    return returnSuccess({"message": "Report endpoint is visible via GET"})

@router.post("/report/")
def test_report_slash(payload: ReportRequest, db: Session = Depends(get_db), org_id: Optional[int] = Header(None)):
    print("DEBUG: test_report_slash (trailing slash) called")
    return get_report(payload, db, org_id)

@router.post("/report")
def get_report(payload: ReportRequest, db: Session = Depends(get_db), org_id: Optional[int] = Header(None)):
    print(f"DEBUG: get_report called with payload={payload}")
    try:
        org_id = org_id or 1
        query = """
        SELECT
            d.dept_name AS department,
            s.semester_desc AS term,
            c.crs_title AS course,
            ma.crs_id AS course_id,
            sec.section AS section,
            ma.section_id AS section_id,
            COUNT(sa.stud_attendance_id) AS absent_count

        FROM lms_manage_attendance ma

        JOIN lms_map_student_attendance sa
            ON sa.attendance_id = ma.attendance_id

        JOIN iems_section sec
            ON sec.id = ma.section_id

        JOIN iems_semester s
            ON s.semester_id = sec.semester_id

        JOIN iems_academic_batch ab
            ON ab.academic_batch_id = s.academic_batch_id

        JOIN iems_program p
            ON p.pgm_id = ab.pgm_id

        JOIN iems_department d
            ON d.dept_id = p.dept_id

        JOIN iems_courses c
            ON c.crs_id = ma.crs_id

        WHERE sa.attendance_status = 'ABSENT'
        AND ma.attendance_date BETWEEN :start AND :end
        AND (ma.org_id = :org OR ma.org_id IS NULL)
        """

        params = {
            "start": payload.start_date,
            "end": payload.end_date,
            "org": org_id
        }

        if payload.department_ids:
            query += " AND d.dept_id IN :dept_ids"
            params["dept_ids"] = payload.department_ids

        if payload.program_ids:
            query += " AND p.pgm_id IN :pgm_ids"
            params["pgm_ids"] = payload.program_ids

        if payload.curriculum_ids:
            query += " AND ab.academic_batch_id IN :batch_ids"
            params["batch_ids"] = payload.curriculum_ids

        if payload.semester_ids:
            query += " AND s.semester_id IN :sem_ids"
            params["sem_ids"] = payload.semester_ids

        if payload.section_ids:
            query += " AND sec.id IN :sec_ids"
            params["sec_ids"] = payload.section_ids

        query += " GROUP BY d.dept_name, s.semester_desc, c.crs_title, ma.crs_id, sec.section, ma.section_id"

        stmt = text(query)
        bind_params = []

        if payload.department_ids:
            bind_params.append(bindparam("dept_ids", expanding=True))

        if payload.program_ids:
            bind_params.append(bindparam("pgm_ids", expanding=True))

        if payload.curriculum_ids:
            bind_params.append(bindparam("batch_ids", expanding=True))

        if payload.semester_ids:
            bind_params.append(bindparam("sem_ids", expanding=True))

        if payload.section_ids:
            bind_params.append(bindparam("sec_ids", expanding=True))

        if bind_params:
            stmt = stmt.bindparams(*bind_params)

        result = db.execute(stmt, params).fetchall()

        return returnSuccess([dict(row._mapping) for row in result])
    except Exception as e:
        return returnException(str(e))

# --------------------------------------------------
# 4. DRILLDOWN
# --------------------------------------------------

@router.post("/drilldown")
def get_drilldown(payload: DrilldownRequest, db: Session = Depends(get_db), org_id: Optional[int] = Header(None)):
    try:
        org_id = org_id or 1
        result = db.execute(text("""
            SELECT
                ma.attendance_date AS date,
                st.name AS student_name,
                st.usno AS usn,
                st.mobile AS student_contact,
                st.fathers_phone AS parent_contact

            FROM lms_manage_attendance ma

            JOIN lms_map_student_attendance sa
                ON sa.attendance_id = ma.attendance_id

            JOIN iems_students st
                ON st.usno = sa.student_usn

            WHERE sa.attendance_status = 'ABSENT'
            AND ma.crs_id = :course_id
            AND ma.section_id = :section_id
            AND ma.attendance_date BETWEEN :start AND :end
            AND (ma.org_id = :org OR ma.org_id IS NULL)

            ORDER BY ma.attendance_date ASC
        """), {
            "course_id": payload.course_id,
            "section_id": payload.section_id,
            "start": payload.start_date,
            "end": payload.end_date,
            "org": org_id
        }).fetchall()

        return returnSuccess([dict(row._mapping) for row in result])
    except Exception as e:
        return returnException(str(e))


# --------------------------------------------------
# 5. EXPORTS
# --------------------------------------------------

from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from fpdf import FPDF

@router.post("/export/excel")
def export_excel(payload: ReportRequest, db: Session = Depends(get_db)):
    data = get_report(payload, db)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Absentees Report"
    
    headers = ["Department", "Term", "Course", "Section", "Absent Count"]
    ws.append(headers)
    
    for row in data:
        ws.append([
            row["department"],
            row["term"],
            row["course"],
            row["section"],
            row["absent_count"]
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=absentees_report.xlsx"}
    )

@router.post("/export/pdf")
def export_pdf(payload: ReportRequest, db: Session = Depends(get_db)):
    data = get_report(payload, db)
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Consolidated Absentees Report", ln=True, align="C")
    pdf.ln(10)
    
    # Table headers
    pdf.set_font("Arial", "B", 10)
    widths = [40, 30, 60, 30, 30]
    headers = ["Dept", "Term", "Course", "Section", "Count"]
    
    for i in range(len(headers)):
        pdf.cell(widths[i], 10, headers[i], border=1)
    pdf.ln()
    
    # Data rows
    pdf.set_font("Arial", "", 9)
    for row in data:
        pdf.cell(widths[0], 10, str(row["department"]), border=1)
        pdf.cell(widths[1], 10, str(row["term"]), border=1)
        pdf.cell(widths[2], 10, str(row["course"]), border=1)
        pdf.cell(widths[3], 10, str(row["section"]), border=1)
        pdf.cell(widths[4], 10, str(row["absent_count"]), border=1)
        pdf.ln()
    
    output = io.BytesIO()
    pdf_content = pdf.output(dest='S').encode('latin-1')
    output.write(pdf_content)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=absentees_report.pdf"}
    )